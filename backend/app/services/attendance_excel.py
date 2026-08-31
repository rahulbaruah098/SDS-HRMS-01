from calendar import monthrange
from copy import copy
from datetime import date, datetime, timedelta
import re
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


GUIDELINE_ROWS = [
    ("Holiday", "H"),
    ("Casual Leave", "CL"),
    ("Casual Leave Half Day", "CLH"),
    ("Earned Leave", "EL"),
    ("Earned Leave Half Day", "ELH"),
    ("Maternity Leave", "ML"),
    ("Paternity Leave", "PL"),
    ("Leave Without Pay", "LWP"),
    ("Leave Without Pay Half Day", "LWPH"),
    ("Compensatory Off", "CO"),
    ("Work From Home", "WFH"),
    ("Tour / Field Work", "T"),
    ("Present", "P"),
    ("Absent", "A"),
]

SUMMARY_COLUMNS = [
    ("CL Availed", "CL"),
    ("EL Availed", "EL"),
    ("LWP", "LWP"),
    ("Half Day", "half_day"),
    ("Remarks", "remarks"),
]

BASE_COLUMNS = [
    "Sl No",
    "Name of Employee",
    "Designation",
    "Project",
    "Location",
    "Emp Code",
]


ORG_FULL_NAMES = {
    "SDS": "Sesta Development Services (SDS)",
    "SDSPL": "Sayanant Development Services Pvt. Ltd. (SDSPL)",
    "SDPL": "Sayanant Development Services Pvt. Ltd. (SDSPL)",
    "AVPL": "Ayanant Ventures Pvt. Ltd. (AVPL)",
    "SDF": "Sayanant Development Foundation (SDF)",
}


THIN_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)

TITLE_FILL = PatternFill("solid", fgColor="92D050")
SUBTITLE_FILL = PatternFill("solid", fgColor="C6E0B4")
HEADER_FILL = PatternFill("solid", fgColor="D9EAD3")
WEEKEND_FILL = PatternFill("solid", fgColor="E2F0D9")
HOLIDAY_FILL = PatternFill("solid", fgColor="FFE699")
ABSENT_FILL = PatternFill("solid", fgColor="F4CCCC")
LEAVE_FILL = PatternFill("solid", fgColor="DDEBF7")
PRESENT_FILL = PatternFill("solid", fgColor="FFFFFF")
SECTION_FILL = PatternFill("solid", fgColor="E2F0D9")
DETAIL_TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
DETAIL_HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
REPORT_TITLE_FILL = PatternFill("solid", fgColor="17365D")
REPORT_META_FILL = PatternFill("solid", fgColor="EAF2F8")
REPORT_ALT_FILL = PatternFill("solid", fgColor="F7FAFC")
MISSING_CHECKOUT_FILL = PatternFill("solid", fgColor="FCE4D6")


def normalize_text(value):
    return str(value or "").strip()


def parse_iso_date(value):
    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    value = normalize_text(value)

    if not value:
        return None

    try:
        return datetime.strptime(value[:10], "%Y-%m-%d").date()
    except Exception:
        return None


def date_key(value):
    parsed = parse_iso_date(value)

    if not parsed:
        return ""

    return parsed.isoformat()


def month_dates(year, month):
    year = int(year)
    month = int(month)
    days = monthrange(year, month)[1]

    return [date(year, month, day) for day in range(1, days + 1)]


def week_dates(start_date, end_date=None):
    start = parse_iso_date(start_date)

    if not start:
        start = date.today()

    end = parse_iso_date(end_date)

    if not end:
        end = start + timedelta(days=6)

    if end < start:
        start, end = end, start

    dates = []
    current = start

    while current <= end:
        dates.append(current)
        current += timedelta(days=1)

    return dates


def day_dates(target_date):
    parsed = parse_iso_date(target_date) or date.today()
    return [parsed]


def year_dates(year):
    year = int(year or date.today().year)

    dates = []
    current = date(year, 1, 1)
    end = date(year, 12, 31)

    while current <= end:
        dates.append(current)
        current += timedelta(days=1)

    return dates


def build_period_dates(period="month", year=None, month=None, date_value=None, week_start=None, week_end=None):
    period = normalize_text(period).lower() or "month"

    if period == "day":
        return day_dates(date_value)

    if period == "week":
        return week_dates(week_start or date_value, week_end)

    if period == "year":
        return year_dates(year)

    today = date.today()
    return month_dates(year or today.year, month or today.month)


def month_name(dates):
    if not dates:
        return ""

    first = dates[0]
    return first.strftime("%B %Y")


def safe_sheet_title(value):
    title = normalize_text(value) or "Attendance"

    invalid_chars = ["\\", "/", "*", "[", "]", ":", "?"]

    for char in invalid_chars:
        title = title.replace(char, "-")

    return title[:31]


def employee_name(employee):
    return (
        normalize_text(employee.get("name"))
        or normalize_text(employee.get("employee_name"))
        or normalize_text(employee.get("full_name"))
        or normalize_text(employee.get("email"))
        or "Employee"
    )


def employee_code(employee):
    return (
        normalize_text(employee.get("emp_code"))
        or normalize_text(employee.get("employee_code"))
        or normalize_text(employee.get("employee_id"))
        or normalize_text(employee.get("code"))
    )


def employee_designation(employee):
    return (
        normalize_text(employee.get("designation"))
        or normalize_text(employee.get("designation_name"))
        or normalize_text(employee.get("title"))
    )


def employee_project(employee):
    return (
        normalize_text(employee.get("project"))
        or normalize_text(employee.get("project_name"))
        or normalize_text(employee.get("department"))
        or "0"
    )


def employee_location(employee):
    return (
        normalize_text(employee.get("location"))
        or normalize_text(employee.get("state"))
        or normalize_text(employee.get("office_state"))
        or normalize_text(employee.get("work_state"))
        or normalize_text(employee.get("branch"))
        or ""
    )


def employee_state(employee):
    return (
        normalize_text(employee.get("state"))
        or normalize_text(employee.get("office_state"))
        or normalize_text(employee.get("work_state"))
        or normalize_text(employee.get("current_state"))
        or normalize_text(employee.get("branch"))
    )


def employee_organisation_name(employee):
    return (
        normalize_text(employee.get("organisation"))
        or normalize_text(employee.get("organization"))
        or normalize_text(employee.get("organisation_name"))
        or normalize_text(employee.get("organization_name"))
        or normalize_text(employee.get("entity"))
    )


def employee_organisation_code(employee):
    return (
        normalize_text(employee.get("organisation_code"))
        or normalize_text(employee.get("organization_code"))
        or normalize_text(employee.get("entity_code"))
    )


def employee_department(employee):
    return (
        normalize_text(employee.get("department"))
        or normalize_text(employee.get("department_name"))
    )


def employee_joining_date(employee):
    return parse_iso_date(
        employee.get("joining_date")
        or employee.get("date_of_joining")
        or employee.get("doj")
    )


def employee_last_working_date(employee):
    return parse_iso_date(
        employee.get("last_working_date")
        or employee.get("resignation_date")
    )


EMPLOYEE_IDENTITY_FIELDS = (
    "employee_id",
    "employee_code",
    "emp_code",
    "code",
)

IDENTITY_MATCH_PRIORITY = (
    "employee_ref",
    "user_ref",
    "alias",
    "email",
)


def normalize_identity_value(value):
    return normalize_text(value).lower()


def looks_like_object_id(value):
    return bool(
        re.fullmatch(
            r"[0-9a-fA-F]{24}",
            normalize_text(value),
        )
    )


def add_unique_identity(target, value):
    value = normalize_identity_value(value)

    if value and value not in target:
        target.append(value)


def stored_identity_aliases(payload):
    payload = payload or {}
    values = payload.get("identity_alias_keys") or []

    if not isinstance(values, (list, tuple, set)):
        return []

    aliases = []

    for value in values:
        add_unique_identity(aliases, value)

    return aliases


def employee_identity_groups(employee):
    """Return employee identifiers in safe matching order.

    Immutable Mongo employee references and linked user references are kept
    separate from editable/legacy employee codes. Phone numbers are deliberately
    excluded because they are not reliable employee identity keys.
    """
    employee = employee or {}
    groups = {
        "employee_ref": [],
        "user_ref": [],
        "alias": [],
        "email": [],
    }

    employee_object_id = employee.get("_id")

    if employee_object_id:
        add_unique_identity(groups["employee_ref"], str(employee_object_id))

    for value in (
        employee.get("employee_ref_id"),
        employee.get("employee_mongo_id"),
        employee.get("employee_record_id"),
    ):
        if looks_like_object_id(value):
            add_unique_identity(groups["employee_ref"], value)
        else:
            add_unique_identity(groups["alias"], value)

    raw_employee_id = employee.get("employee_id")

    if looks_like_object_id(raw_employee_id):
        add_unique_identity(groups["employee_ref"], raw_employee_id)
    else:
        add_unique_identity(groups["alias"], raw_employee_id)

    raw_id = employee.get("id")

    if looks_like_object_id(raw_id):
        add_unique_identity(groups["employee_ref"], raw_id)
    else:
        add_unique_identity(groups["alias"], raw_id)

    add_unique_identity(groups["user_ref"], employee.get("user_id"))

    for alias in stored_identity_aliases(employee):
        add_unique_identity(groups["alias"], alias)

    for field_name in EMPLOYEE_IDENTITY_FIELDS:
        value = employee.get(field_name)

        if field_name == "employee_id" and looks_like_object_id(value):
            continue

        add_unique_identity(groups["alias"], value)

    add_unique_identity(groups["email"], employee.get("email"))
    add_unique_identity(groups["email"], employee.get("official_email"))

    return groups


def attendance_row_identity_groups(row):
    """Return attendance/leave identifiers without using the row's own _id."""
    row = row or {}
    groups = {
        "employee_ref": [],
        "user_ref": [],
        "alias": [],
        "email": [],
    }

    for value in (
        row.get("employee_ref_id"),
        row.get("employee_mongo_id"),
        row.get("employee_record_id"),
    ):
        if looks_like_object_id(value):
            add_unique_identity(groups["employee_ref"], value)
        else:
            add_unique_identity(groups["alias"], value)

    raw_employee_id = row.get("employee_id")

    if looks_like_object_id(raw_employee_id):
        add_unique_identity(groups["employee_ref"], raw_employee_id)
    else:
        add_unique_identity(groups["alias"], raw_employee_id)

    add_unique_identity(groups["user_ref"], row.get("user_id"))

    for alias in stored_identity_aliases(row):
        add_unique_identity(groups["alias"], alias)

    for field_name in EMPLOYEE_IDENTITY_FIELDS:
        value = row.get(field_name)

        if field_name == "employee_id" and looks_like_object_id(value):
            continue

        add_unique_identity(groups["alias"], value)

    add_unique_identity(groups["email"], row.get("email"))
    add_unique_identity(groups["email"], row.get("official_email"))

    return groups


def flatten_identity_groups(groups):
    values = []

    for group_name in IDENTITY_MATCH_PRIORITY:
        for value in groups.get(group_name, []):
            add_unique_identity(values, value)

    return values


def employee_identifier_values(employee):
    return flatten_identity_groups(employee_identity_groups(employee))


def attendance_row_identifier_values(row):
    return flatten_identity_groups(attendance_row_identity_groups(row))


def attendance_employee_identifier(row):
    identifiers = attendance_row_identifier_values(row)
    return identifiers[0] if identifiers else ""


def identity_tenant_id(payload):
    return normalize_identity_value((payload or {}).get("tenant_id"))


def ambiguous_fallback_identities(employees):
    """Find editable aliases/emails shared by multiple employee masters.

    Such values are excluded from fallback matching so an old duplicate code or
    shared email cannot silently place attendance against the wrong employee.
    """
    owners_by_key = {
        "alias": {},
        "email": {},
    }

    for index, employee in enumerate(employees or []):
        groups = employee_identity_groups(employee)
        tenant_id = identity_tenant_id(employee)
        owner = (
            groups["employee_ref"][0]
            if groups["employee_ref"]
            else groups["user_ref"][0]
            if groups["user_ref"]
            else f"employee-row-{index}"
        )

        for group_name in ("alias", "email"):
            for value in groups[group_name]:
                for tenant_key in {tenant_id, ""}:
                    key = (tenant_key, value)
                    owners_by_key[group_name].setdefault(key, set()).add(owner)

    return {
        group_name: {
            key
            for key, owners in owners_by_key[group_name].items()
            if len(owners) > 1
        }
        for group_name in ("alias", "email")
    }


def build_employee_identity_index(employees):
    """Build a tenant-safe identity index for attendance detail rows.

    A key that belongs to more than one employee is marked ambiguous and is
    never used.  Immutable employee references are tried before user links,
    editable codes, or email addresses.
    """
    index = {
        group_name: {}
        for group_name in IDENTITY_MATCH_PRIORITY
    }

    for employee in employees or []:
        groups = employee_identity_groups(employee)
        tenant_id = identity_tenant_id(employee)

        for group_name in IDENTITY_MATCH_PRIORITY:
            for value in groups[group_name]:
                # The tenantless key is safe here because every report route
                # passes an already tenant-scoped employee list.  If the same
                # key appears twice it is deliberately made unusable.
                for tenant_key in {tenant_id, ""}:
                    key = (tenant_key, value)
                    current = index[group_name].get(key)

                    if current is None and key in index[group_name]:
                        continue

                    if current is not None and current is not employee:
                        index[group_name][key] = None
                    else:
                        index[group_name][key] = employee

    return index


def matching_employee_for_row(row, employees=None, identity_index=None):
    """Resolve one attendance/leave row to exactly one employee master."""
    identity_index = identity_index or build_employee_identity_index(employees)
    groups = attendance_row_identity_groups(row)
    tenant_id = identity_tenant_id(row)
    tenant_candidates = [tenant_id]

    if tenant_id:
        tenant_candidates.append("")

    for group_name in IDENTITY_MATCH_PRIORITY:
        for value in groups[group_name]:
            for tenant_key in tenant_candidates:
                employee = identity_index.get(group_name, {}).get(
                    (tenant_key, value)
                )

                if employee is not None:
                    return employee

    return None


def empty_identity_lookup():
    return {
        group_name: {}
        for group_name in IDENTITY_MATCH_PRIORITY
    }


def add_identity_lookup_row(
    lookup,
    row,
    day_key,
    code,
    ambiguous_identities=None,
):
    groups = attendance_row_identity_groups(row)
    tenant_id = identity_tenant_id(row)
    ambiguous_identities = ambiguous_identities or {
        "alias": set(),
        "email": set(),
    }

    for group_name in IDENTITY_MATCH_PRIORITY:
        for value in groups[group_name]:
            if (
                group_name in {"alias", "email"}
                and (tenant_id, value) in ambiguous_identities.get(group_name, set())
            ):
                continue

            lookup[group_name][(tenant_id, value, day_key)] = code


def identity_lookup_code(lookup, employee, day_key):
    groups = employee_identity_groups(employee)
    tenant_id = identity_tenant_id(employee)
    tenant_candidates = [tenant_id]

    # Tenantless historical rows may still be used only after exact-tenant
    # matching fails. Cross-tenant identifiers are never searched.
    if tenant_id:
        tenant_candidates.append("")

    for group_name in IDENTITY_MATCH_PRIORITY:
        for value in groups[group_name]:
            for tenant_key in tenant_candidates:
                code = lookup.get(group_name, {}).get(
                    (tenant_key, value, day_key)
                )

                if code:
                    return code

    return ""


def attendance_status_code(row):
    status = normalize_text(row.get("status")).lower()
    mode = normalize_text(row.get("mode")).lower()

    if mode in {"wfh", "work_from_home", "work from home"}:
        return "WFH"

    if mode in {"field", "tour", "travel", "official_tour", "official tour"}:
        return "T"

    if status in {"present", "checked_in", "checked-out", "checked_out", "verified", "approved"}:
        return "P"

    if row.get("check_in") or row.get("check_in_time") or row.get("checked_in_at"):
        return "P"

    return ""

def normalize_leave_type_for_excel(value):
    value = normalize_text(value).upper()

    aliases = {
        "CL": "CL",
        "CASUAL": "CL",
        "CASUAL LEAVE": "CL",
        "CASUAL_LEAVE": "CL",

        "EL": "EL",
        "EARNED": "EL",
        "EARNED LEAVE": "EL",
        "EARNED_LEAVE": "EL",

        "COMP OFF": "CO",
        "COMP-OFF": "CO",
        "COMPOFF": "CO",
        "COMPENSATORY LEAVE": "CO",
        "COMPENSATORY OFF": "CO",

        "HALF DAY": "HALF-DAY",
        "HALF-DAY": "HALF-DAY",
        "HALFDAY": "HALF-DAY",
        "HD": "HALF-DAY",

        "LWP": "LWP",
        "LEAVE WITHOUT PAY": "LWP",
        "LOSS OF PAY": "LWP",

        "ML": "ML",
        "MATERNITY": "ML",
        "MATERNITY LEAVE": "ML",

        "PL": "PL",
        "PATERNITY": "PL",
        "PATERNITY LEAVE": "PL",
    }

    return aliases.get(value, value)


def is_half_day_leave(row):
    requested_type = normalize_leave_type_for_excel(
        row.get("requested_leave_type")
        or row.get("requested_leave_type_label")
        or row.get("leave_type")
        or row.get("leave_type_label")
        or row.get("type")
    )

    if requested_type == "HALF-DAY":
        return True

    day_type = normalize_text(
        row.get("day_type")
        or row.get("duration_type")
        or row.get("leave_duration")
        or row.get("duration")
    ).lower().replace("-", "_").replace(" ", "_")

    if day_type in {
        "half_day",
        "half",
        "first_half",
        "second_half",
        "morning_half",
        "afternoon_half",
    }:
        return True

    session_value = normalize_text(
        row.get("half_day_session")
        or row.get("session")
        or row.get("leave_session")
    ).lower().replace("-", "_").replace(" ", "_")

    if session_value in {
        "first_half",
        "second_half",
        "morning",
        "afternoon",
        "forenoon",
        "afternoon_session",
    }:
        return True

    if str(row.get("half_day") or row.get("is_half_day") or "").lower() in {
        "true",
        "1",
        "yes",
        "on",
    }:
        return True

    for key in ("leave_days", "days", "total_days", "approved_days"):
        try:
            if float(row.get(key, 0) or 0) == 0.5:
                return True
        except Exception:
            pass

    for key in ("lwp_days", "deducted_days"):
        try:
            if float(row.get(key, 0) or 0) == 0.5:
                return True
        except Exception:
            pass

    return False


def leave_status_code(row):
    status = normalize_text(row.get("status")).lower()
    approval_stage = normalize_text(row.get("approval_stage")).lower()
    final_status = normalize_text(row.get("final_status")).lower()

    approved_values = {
        "approved",
        "accepted",
        "hr_approved",
        "manager_approved",
        "ro_approved",
    }

    if (
        status not in approved_values
        and approval_stage not in approved_values
        and final_status not in approved_values
    ):
        return ""

    requested_leave_type = normalize_leave_type_for_excel(
        row.get("requested_leave_type")
        or row.get("requested_leave_type_label")
        or row.get("leave_type")
        or row.get("leave_type_label")
        or row.get("type")
    )

    deducted_leave_type = normalize_leave_type_for_excel(
        row.get("deducted_leave_type")
        or row.get("deducted_leave_type_label")
        or row.get("approved_leave_type")
        or row.get("approved_leave_type_label")
        or ""
    )

    leave_type = deducted_leave_type or requested_leave_type
    half_day = is_half_day_leave(row)

    try:
        lwp_days = float(row.get("lwp_days", 0) or 0)
    except Exception:
        lwp_days = 0

    if requested_leave_type == "HALF-DAY":
        if deducted_leave_type == "EL":
            return "ELH"

        if deducted_leave_type == "LWP" or lwp_days > 0:
            return "LWPH"

        if deducted_leave_type == "CL":
            return "CLH"

        return "CLH"

    if leave_type == "CL":
        return "CLH" if half_day else "CL"

    if leave_type == "EL":
        return "ELH" if half_day else "EL"

    if leave_type == "LWP":
        return "LWPH" if half_day else "LWP"

    if leave_type == "ML":
        return "ML"

    if leave_type == "PL":
        return "PL"

    if leave_type == "CO":
        return "CO"

    return leave_type or "L"


def leave_date_range(row):
    start = parse_iso_date(
        row.get("from_date")
        or row.get("start_date")
        or row.get("date")
    )

    end = parse_iso_date(
        row.get("to_date")
        or row.get("upto_date")
        or row.get("end_date")
        or row.get("date")
    )

    if not start and end:
        start = end

    if not end and start:
        end = start

    return start, end


def build_attendance_lookup(attendance_logs, employees=None):
    lookup = empty_identity_lookup()
    ambiguous_identities = ambiguous_fallback_identities(employees)

    for row in attendance_logs or []:
        employee_keys = attendance_row_identifier_values(row)
        day_key = date_key(
            row.get("date")
            or row.get("attendance_date")
            or row.get("attendance_day")
            or row.get("created_at")
            or row.get("check_in")
            or row.get("check_in_time")
            or row.get("checked_in_at")
        )

        if not employee_keys or not day_key:
            continue

        code = attendance_status_code(row)

        if not code:
            continue

        add_identity_lookup_row(
            lookup,
            row,
            day_key,
            code,
            ambiguous_identities=ambiguous_identities,
        )

    return lookup


def build_leave_lookup(leave_requests, employees=None):
    lookup = empty_identity_lookup()
    ambiguous_identities = ambiguous_fallback_identities(employees)

    for row in leave_requests or []:
        code = leave_status_code(row)

        if not code:
            continue

        employee_keys = attendance_row_identifier_values(row)
        start, end = leave_date_range(row)

        if not employee_keys or not start or not end:
            continue

        current = start

        while current <= end:
            add_identity_lookup_row(
                lookup,
                row,
                current.isoformat(),
                code,
                ambiguous_identities=ambiguous_identities,
            )
            current += timedelta(days=1)

    return lookup


def build_holiday_lookup(holidays):
    lookup = {}

    for row in holidays or []:
        day_key = date_key(row.get("date") or row.get("holiday_date"))

        if not day_key:
            continue

        title = normalize_text(row.get("title") or row.get("name") or "Holiday")
        state_key = normalize_identity_value(
            row.get("state")
            or row.get("office_state")
            or row.get("work_state")
            or row.get("branch")
        )

        lookup[(state_key, day_key)] = title

        # Retain the legacy date-only key for callers that supply a single
        # state-specific holiday list.
        lookup.setdefault(day_key, title)

    return lookup


def holiday_for_employee_date(employee, target_date, holiday_lookup):
    target_key = target_date.isoformat()
    state_key = normalize_identity_value(employee_state(employee))

    if state_key and (state_key, target_key) in holiday_lookup:
        return holiday_lookup[(state_key, target_key)]

    if ("", target_key) in holiday_lookup:
        return holiday_lookup[("", target_key)]

    # Date-only entries are accepted only when there are no state-specific
    # entries for that day.  This prevents a holiday from one state making all
    # other states absent/closed in an all-state report.
    has_state_specific_entry = any(
        isinstance(key, tuple) and len(key) == 2 and key[1] == target_key and key[0]
        for key in holiday_lookup
    )

    if not has_state_specific_entry:
        return holiday_lookup.get(target_key, "")

    return ""


def is_standard_weekly_off(target_date):
    if target_date.weekday() == 6:
        return True

    if target_date.weekday() != 5:
        return False

    saturday_number = ((target_date.day - 1) // 7) + 1
    return saturday_number in {2, 4}


def code_for_employee_date(
    employee,
    target_date,
    attendance_lookup,
    leave_lookup,
    holiday_lookup,
    today_reference=None,
):
    target_key = target_date.isoformat()
    today_reference = today_reference or date.today()

    joining_date = employee_joining_date(employee)
    last_working_date = employee_last_working_date(employee)

    if target_date > today_reference:
        return ""

    if joining_date and target_date < joining_date:
        return ""

    if last_working_date and target_date > last_working_date:
        return ""

    # Real attendance has priority over leave because an employee who actually
    # checked in must never be exported as absent or only-on-leave.
    attendance_code = identity_lookup_code(
        attendance_lookup,
        employee,
        target_key,
    )

    if attendance_code:
        return attendance_code

    leave_code = identity_lookup_code(
        leave_lookup,
        employee,
        target_key,
    )

    if leave_code:
        return leave_code

    if holiday_for_employee_date(employee, target_date, holiday_lookup):
        return "H"

    if is_standard_weekly_off(target_date):
        return ""

    return "A"


def count_code(row_codes, target_code):
    return sum(1 for code in row_codes if normalize_text(code).upper() == target_code)


def count_half_day_codes(row_codes):
    half_day_codes = {"CLH", "ELH", "LWPH"}

    return sum(1 for code in row_codes if normalize_text(code).upper() in half_day_codes)


def apply_cell_style(cell, fill=None, bold=False, size=11, align="center", vertical="center"):
    cell.font = Font(name="Calibri", size=size, bold=bold, color="000000")
    cell.alignment = Alignment(horizontal=align, vertical=vertical, wrap_text=True)
    cell.border = THIN_BORDER

    if fill:
        cell.fill = fill


def style_status_cell(cell, value):
    code = normalize_text(value).upper()

    if code == "P":
        fill = PRESENT_FILL
    elif code in {"CL", "CLH", "EL", "ELH", "ML", "PL", "LWP", "LWPH", "CO"}:
        fill = LEAVE_FILL
    elif code == "H":
        fill = HOLIDAY_FILL
    elif code == "A":
        fill = ABSENT_FILL
    elif code in {"WFH", "T"}:
        fill = WEEKEND_FILL
    else:
        fill = PRESENT_FILL

    apply_cell_style(cell, fill=fill)


def clean_excel_value(value):
    if isinstance(value, datetime):
        return value.strftime("%d-%m-%Y %I:%M %p")

    if isinstance(value, date):
        return value.strftime("%d-%m-%Y")

    if value is None:
        return ""

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return value

    return str(value)


def location_text(row, key="check_in_location"):
    location = row.get(key) or row.get("location") or {}

    if not isinstance(location, dict):
        return ""

    lat = location.get("latitude") or location.get("lat")
    lng = location.get("longitude") or location.get("lng")
    accuracy = location.get("accuracy")

    if not lat or not lng:
        return ""

    text = f"{lat}, {lng}"

    if accuracy:
        try:
            text = f"{text} ±{round(float(accuracy))}m"
        except Exception:
            text = f"{text} ±{accuracy}m"

    return text


def location_map_url(row, key="check_in_location"):
    location = row.get(key) or row.get("location") or {}

    if not isinstance(location, dict):
        return ""

    lat = location.get("latitude") or location.get("lat")
    lng = location.get("longitude") or location.get("lng")

    if not lat or not lng:
        return ""

    return f"https://www.google.com/maps?q={lat},{lng}"


def add_detail_sheet(wb, title, columns, rows):
    ws = wb.create_sheet(safe_sheet_title(title))

    final_col = len(columns)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=final_col)
    ws.cell(1, 1).value = title
    ws.cell(1, 1).font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws.cell(1, 1).fill = DETAIL_TITLE_FILL
    ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(1, 1).border = THIN_BORDER

    for col in range(1, final_col + 1):
        cell = ws.cell(1, col)
        cell.fill = DETAIL_TITLE_FILL
        cell.border = THIN_BORDER

    for col_index, (label, _) in enumerate(columns, start=1):
        cell = ws.cell(3, col_index)
        cell.value = label
        apply_cell_style(cell, fill=DETAIL_HEADER_FILL, bold=True)

    start_row = 4

    if not rows:
        ws.cell(start_row, 1).value = "No records found."
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=final_col)
        apply_cell_style(ws.cell(start_row, 1), align="center", bold=True)
    else:
        for row_index, row in enumerate(rows, start=start_row):
            for col_index, (_, key) in enumerate(columns, start=1):
                cell = ws.cell(row_index, col_index)
                cell.value = clean_excel_value(row.get(key))
                apply_cell_style(cell, align="left")

    for col_index, (label, key) in enumerate(columns, start=1):
        max_len = len(label)

        for row in rows or []:
            max_len = max(max_len, len(str(clean_excel_value(row.get(key)))))

        ws.column_dimensions[get_column_letter(col_index)].width = min(max(max_len + 3, 14), 45)

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(final_col)}{max(start_row, start_row + len(rows or []) - 1)}"

    page_setup = ws.page_setup
    page_setup.orientation = "landscape"
    page_setup.fitToWidth = 1
    page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    return ws


def add_professional_table_sheet(
    wb,
    title,
    columns,
    rows,
    metadata=None,
    sheet_title=None,
    status_key="status",
):
    """Create a clean, filterable report sheet with report metadata."""
    if (
        wb.sheetnames == ["Sheet"]
        and wb.active.max_row == 1
        and wb.active.max_column == 1
        and wb.active["A1"].value is None
    ):
        ws = wb.active
        ws.title = safe_sheet_title(sheet_title or title)
    else:
        ws = wb.create_sheet(safe_sheet_title(sheet_title or title))

    rows = list(rows or [])
    metadata = [normalize_text(value) for value in (metadata or []) if normalize_text(value)]
    final_col = max(len(columns), 1)
    final_col_letter = get_column_letter(final_col)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=final_col)
    ws.cell(1, 1).value = title
    ws.cell(1, 1).font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    ws.cell(1, 1).fill = REPORT_TITLE_FILL
    ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")

    for column in range(1, final_col + 1):
        ws.cell(1, column).fill = REPORT_TITLE_FILL
        ws.cell(1, column).border = THIN_BORDER

    meta_start_row = 2

    for offset, value in enumerate(metadata):
        row_number = meta_start_row + offset
        ws.merge_cells(
            start_row=row_number,
            start_column=1,
            end_row=row_number,
            end_column=final_col,
        )
        ws.cell(row_number, 1).value = value
        ws.cell(row_number, 1).font = Font(name="Calibri", size=10, bold=False, color="1F2937")
        ws.cell(row_number, 1).fill = REPORT_META_FILL
        ws.cell(row_number, 1).alignment = Alignment(horizontal="left", vertical="center")

        for column in range(1, final_col + 1):
            ws.cell(row_number, column).fill = REPORT_META_FILL
            ws.cell(row_number, column).border = THIN_BORDER

    header_row = max(4, meta_start_row + len(metadata) + 1)

    for column_index, (label, _) in enumerate(columns, start=1):
        cell = ws.cell(header_row, column_index)
        cell.value = label
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = DETAIL_TITLE_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    data_start_row = header_row + 1

    if not rows:
        ws.merge_cells(
            start_row=data_start_row,
            start_column=1,
            end_row=data_start_row,
            end_column=final_col,
        )
        ws.cell(data_start_row, 1).value = "No records found for the selected filters."
        apply_cell_style(ws.cell(data_start_row, 1), align="center", bold=True)
    else:
        for row_index, row in enumerate(rows, start=data_start_row):
            status = normalize_text(row.get(status_key)).lower()

            for column_index, (_, key) in enumerate(columns, start=1):
                cell = ws.cell(row_index, column_index)
                cell.value = clean_excel_value(row.get(key))
                apply_cell_style(cell, align="left")

                if status == "absent":
                    cell.fill = ABSENT_FILL
                elif status in {"not checked out", "not_checked_out", "missing checkout"}:
                    cell.fill = MISSING_CHECKOUT_FILL
                elif (row_index - data_start_row) % 2:
                    cell.fill = REPORT_ALT_FILL

    for column_index, (label, key) in enumerate(columns, start=1):
        max_length = len(label)

        for row in rows:
            max_length = max(
                max_length,
                len(str(clean_excel_value(row.get(key)))),
            )

        ws.column_dimensions[get_column_letter(column_index)].width = min(
            max(max_length + 3, 12),
            42,
        )

    final_data_row = max(data_start_row, data_start_row + len(rows) - 1)
    ws.freeze_panes = f"A{data_start_row}"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[header_row].height = 30
    ws.auto_filter.ref = f"A{header_row}:{final_col_letter}{final_data_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = f"1:{header_row}"

    return ws


def report_period_label(dates, period="month"):
    dates = list(dates or [])

    if not dates:
        return "Selected period"

    period_key = normalize_text(period).lower()

    if period_key == "day":
        return dates[0].strftime("%d %B %Y")

    if period_key == "week":
        return (
            f"{dates[0].strftime('%d %B %Y')} to "
            f"{dates[-1].strftime('%d %B %Y')}"
        )

    if period_key == "year":
        return str(dates[0].year)

    return dates[0].strftime("%B %Y")


def employee_report_identity(employee):
    groups = employee_identity_groups(employee)

    for group_name in IDENTITY_MATCH_PRIORITY:
        if groups[group_name]:
            return groups[group_name][0]

    return employee_name(employee).lower()


def attendance_check_in_value(row):
    return (
        row.get("check_in")
        or row.get("check_in_at")
        or row.get("check_in_time")
        or row.get("checked_in_at")
    )


def attendance_check_out_value(row):
    return (
        row.get("check_out")
        or row.get("check_out_at")
        or row.get("check_out_time")
        or row.get("checked_out_at")
    )


def exception_employee_fields(employee, source_row=None):
    employee = employee or {}
    source_row = source_row or {}

    return {
        "employee_id": str(employee.get("_id") or source_row.get("employee_id") or ""),
        "employee_code": employee_code(employee) or normalize_text(
            source_row.get("employee_code")
            or source_row.get("emp_code")
            or source_row.get("code")
        ),
        "employee_name": employee_name(employee) if employee else (
            normalize_text(source_row.get("employee_name"))
            or normalize_text(source_row.get("name"))
            or "Employee"
        ),
        "organisation_code": employee_organisation_code(employee) or normalize_text(
            source_row.get("organisation_code")
            or source_row.get("organization_code")
        ),
        "organisation": employee_organisation_name(employee) or normalize_text(
            source_row.get("organisation")
            or source_row.get("organization")
            or source_row.get("organisation_name")
            or source_row.get("organization_name")
        ),
        "state": employee_state(employee) or normalize_text(
            source_row.get("state")
            or source_row.get("office_state")
            or source_row.get("work_state")
            or source_row.get("branch")
        ),
        "department": employee_department(employee) or normalize_text(
            source_row.get("department") or source_row.get("department_name")
        ),
        "designation": employee_designation(employee) or normalize_text(
            source_row.get("designation") or source_row.get("designation_name")
        ),
        "team_leader": normalize_text(
            employee.get("team_leader_name") or source_row.get("team_leader_name")
        ),
        "reporting_officer": normalize_text(
            employee.get("reporting_officer_name")
            or source_row.get("reporting_officer_name")
        ),
    }


def build_attendance_exception_rows(
    employees=None,
    attendance_logs=None,
    leave_requests=None,
    holidays=None,
    dates=None,
    today_reference=None,
):
    """Return day-level absent and missing-checkout rows for HR reports."""
    employees = list(employees or [])
    attendance_logs = list(attendance_logs or [])
    dates = sorted(set(dates or []))
    today_reference = today_reference or date.today()

    attendance_lookup = build_attendance_lookup(
        attendance_logs,
        employees=employees,
    )
    leave_lookup = build_leave_lookup(
        leave_requests or [],
        employees=employees,
    )
    holiday_lookup = build_holiday_lookup(holidays or [])
    identity_index = build_employee_identity_index(employees)

    absent_rows = []

    for employee in sorted(
        employees,
        key=lambda item: (
            employee_organisation_code(item).lower(),
            employee_state(item).lower(),
            employee_name(item).lower(),
        ),
    ):
        base_fields = exception_employee_fields(employee)

        for target_date in dates:
            code = code_for_employee_date(
                employee,
                target_date,
                attendance_lookup,
                leave_lookup,
                holiday_lookup,
                today_reference=today_reference,
            )

            if code != "A":
                continue

            absent_rows.append({
                **base_fields,
                "date": target_date.isoformat(),
                "attendance_code": "A",
                "status": "Absent",
                "remarks": "No attendance, approved leave, holiday, or weekly off found.",
            })

    selected_date_keys = {item.isoformat() for item in dates}
    grouped_logs = {}

    for row in attendance_logs:
        row_date = date_key(
            row.get("date")
            or row.get("attendance_date")
            or attendance_check_in_value(row)
        )

        if not row_date or (selected_date_keys and row_date not in selected_date_keys):
            continue

        if not attendance_check_in_value(row):
            continue

        matched_employee = matching_employee_for_row(
            row,
            employees=employees,
            identity_index=identity_index,
        )

        # When an employee filter is active, unmatched logs must not leak into
        # the selected employee's report.
        if employees and matched_employee is None:
            continue

        identity_key = (
            employee_report_identity(matched_employee)
            if matched_employee
            else attendance_employee_identifier(row)
        )
        group_key = (identity_key, row_date)
        grouped_logs.setdefault(group_key, []).append((row, matched_employee))

    missing_checkout_rows = []

    for (_, row_date), grouped_rows in grouped_logs.items():
        if any(attendance_check_out_value(row) for row, _ in grouped_rows):
            continue

        selected_row, matched_employee = max(
            grouped_rows,
            key=lambda item: clean_excel_value(attendance_check_in_value(item[0])),
        )
        base_fields = exception_employee_fields(matched_employee, selected_row)

        missing_checkout_rows.append({
            **base_fields,
            "date": row_date,
            "mode": normalize_text(selected_row.get("mode") or "office"),
            "check_in": clean_excel_value(attendance_check_in_value(selected_row)),
            "check_out": "",
            "check_in_location": location_text(selected_row, "check_in_location"),
            "check_in_map_url": location_map_url(selected_row, "check_in_location"),
            "late_reason": normalize_text(selected_row.get("late_reason")),
            "status": "Not Checked Out",
            "remarks": "Checked in, but no checkout was recorded.",
        })

    absent_rows.sort(key=lambda row: (row["date"], row["employee_name"].lower()))
    missing_checkout_rows.sort(
        key=lambda row: (row["date"], row["employee_name"].lower())
    )

    for index, row in enumerate(absent_rows, start=1):
        row["sl_no"] = index

    for index, row in enumerate(missing_checkout_rows, start=1):
        row["sl_no"] = index

    return {
        "absent": absent_rows,
        "missing_checkout": missing_checkout_rows,
        "summary": {
            "absent_records": len(absent_rows),
            "absent_employees": len({row["employee_id"] for row in absent_rows}),
            "missing_checkout_records": len(missing_checkout_rows),
            "missing_checkout_employees": len({
                row["employee_id"] for row in missing_checkout_rows
            }),
        },
    }


ABSENT_REPORT_COLUMNS = [
    ("Sl No", "sl_no"),
    ("Employee Code", "employee_code"),
    ("Employee Name", "employee_name"),
    ("Organisation Code", "organisation_code"),
    ("Organisation / Entity", "organisation"),
    ("State", "state"),
    ("Department", "department"),
    ("Designation", "designation"),
    ("Team Leader", "team_leader"),
    ("Reporting Officer", "reporting_officer"),
    ("Absent Date", "date"),
    ("Attendance Code", "attendance_code"),
    ("Status", "status"),
    ("Remarks", "remarks"),
]


MISSING_CHECKOUT_REPORT_COLUMNS = [
    ("Sl No", "sl_no"),
    ("Employee Code", "employee_code"),
    ("Employee Name", "employee_name"),
    ("Organisation Code", "organisation_code"),
    ("Organisation / Entity", "organisation"),
    ("State", "state"),
    ("Department", "department"),
    ("Designation", "designation"),
    ("Team Leader", "team_leader"),
    ("Reporting Officer", "reporting_officer"),
    ("Attendance Date", "date"),
    ("Mode", "mode"),
    ("Check In", "check_in"),
    ("Check Out", "check_out"),
    ("Check-In Location", "check_in_location"),
    ("Check-In Map", "check_in_map_url"),
    ("Late Reason", "late_reason"),
    ("Status", "status"),
    ("Remarks", "remarks"),
]

def field_attendance_detail_rows(attendance_logs):
    rows = []

    for index, row in enumerate(attendance_logs or [], start=1):
        mode = normalize_text(row.get("mode")).lower()

        if mode not in {"field", "tour", "travel", "official_tour", "official tour"}:
            continue

        rows.append({
            "sl_no": index,
            "employee_name": (
                normalize_text(row.get("employee_name"))
                or normalize_text(row.get("name"))
                or "Employee"
            ),
            "employee_code": (
                normalize_text(row.get("employee_code"))
                or normalize_text(row.get("emp_code"))
                or normalize_text(row.get("employee_id"))
            ),
            "department": normalize_text(row.get("department")),
            "designation": normalize_text(row.get("designation")),
            "team_leader": normalize_text(row.get("team_leader_name")),
            "reporting_officer": normalize_text(row.get("reporting_officer_name")),
            "date": normalize_text(row.get("date")),
            "mode": normalize_text(row.get("mode")),
            "field_location": normalize_text(row.get("field_location")),
            "field_photo_url": (
                normalize_text(row.get("field_photo"))
                or normalize_text(row.get("proof_photo"))
                or normalize_text(row.get("photo"))
            ),
            "check_in": clean_excel_value(row.get("check_in")),
            "check_out": clean_excel_value(row.get("check_out")),
            "check_in_location": location_text(row, "check_in_location"),
            "check_in_map_url": location_map_url(row, "check_in_location"),
            "check_out_location": location_text(row, "check_out_location"),
            "check_out_map_url": location_map_url(row, "check_out_location"),
            "holiday_work_approval": (
                normalize_text(row.get("holiday_work_approval_status"))
                or normalize_text(row.get("holiday_work_status"))
                or ("approved" if row.get("holiday_work_request_id") else "")
            ),
            "holiday_title": normalize_text(row.get("holiday_title") or row.get("holiday_name")),
            "verified_by": (
                normalize_text(row.get("verified_by_name"))
                or normalize_text(row.get("approved_by_name"))
                or normalize_text(row.get("decided_by_name"))
            ),
            "status": normalize_text(row.get("status")),
        })

    return rows


def holiday_work_detail_rows(holiday_work_requests):
    rows = []

    for index, row in enumerate(holiday_work_requests or [], start=1):
        rows.append({
            "sl_no": index,
            "employee_name": (
                normalize_text(row.get("employee_name"))
                or normalize_text(row.get("name"))
                or "Employee"
            ),
            "employee_code": (
                normalize_text(row.get("employee_code"))
                or normalize_text(row.get("emp_code"))
                or normalize_text(row.get("employee_id"))
            ),
            "department": normalize_text(row.get("department")),
            "designation": normalize_text(row.get("designation")),
            "team_leader": normalize_text(row.get("team_leader_name")),
            "reporting_officer": normalize_text(row.get("reporting_officer_name")),
            "date": normalize_text(row.get("date")),
            "holiday_title": normalize_text(row.get("holiday_title") or row.get("holiday_name")),
            "holiday_type": normalize_text(row.get("holiday_type")),
            "reason": normalize_text(row.get("reason")),
            "work_location": normalize_text(row.get("work_location") or row.get("field_location")),
            "proof_photo_url": (
                normalize_text(row.get("proof_photo"))
                or normalize_text(row.get("field_photo"))
                or normalize_text(row.get("photo"))
            ),
            "location": location_text(row, "location"),
            "map_url": location_map_url(row, "location"),
            "approval_stage": normalize_text(row.get("approval_stage")),
            "status": normalize_text(row.get("status")),
            "decided_by": (
                normalize_text(row.get("decided_by_name"))
                or normalize_text(row.get("approved_by_name"))
                or normalize_text(row.get("rejected_by_name"))
            ),
            "decided_at": clean_excel_value(
                row.get("decided_at")
                or row.get("approved_at")
                or row.get("rejected_at")
            ),
            "created_at": clean_excel_value(row.get("created_at")),
        })

    return rows

def compoff_detail_rows(compoff_credits):
    rows = []

    for index, row in enumerate(compoff_credits or [], start=1):
        rows.append({
            "sl_no": index,
            "employee_name": (
                normalize_text(row.get("employee_name"))
                or normalize_text(row.get("name"))
                or "Employee"
            ),
            "employee_code": (
                normalize_text(row.get("employee_code"))
                or normalize_text(row.get("emp_code"))
                or normalize_text(row.get("employee_id"))
            ),
            "department": normalize_text(row.get("department")),
            "designation": normalize_text(row.get("designation")),
            "team_leader": normalize_text(row.get("team_leader_name")),
            "reporting_officer": normalize_text(row.get("reporting_officer_name")),
            "earned_date": normalize_text(row.get("earned_date")),
            "claim_from_date": normalize_text(row.get("claim_from_date") or row.get("available_from")),
            "expiry_date": normalize_text(row.get("expiry_date") or row.get("valid_until")),
            "claim_date": normalize_text(row.get("claim_date") or row.get("claimed_date")),
            "holiday_title": normalize_text(row.get("holiday_title") or row.get("holiday_name")),
            "holiday_work_request_id": normalize_text(row.get("holiday_work_request_id")),
            "attendance_log_id": normalize_text(row.get("attendance_log_id")),
            "leave_request_id": normalize_text(row.get("leave_request_id")),
            "status": normalize_text(row.get("status")),
        })

    return rows

def create_field_attendance_detail_sheet(wb, attendance_logs):
    columns = [
        ("Sl No", "sl_no"),
        ("Employee Name", "employee_name"),
        ("Emp Code", "employee_code"),
        ("Department", "department"),
        ("Designation", "designation"),
        ("Team Leader", "team_leader"),
        ("Reporting Officer", "reporting_officer"),
        ("Date", "date"),
        ("Mode", "mode"),
        ("Field Place", "field_location"),
        ("Field Photo URL", "field_photo_url"),
        ("Check In", "check_in"),
        ("Check Out", "check_out"),
        ("Check-In Location", "check_in_location"),
        ("Check-In Map Link", "check_in_map_url"),
        ("Check-Out Location", "check_out_location"),
        ("Check-Out Map Link", "check_out_map_url"),
        ("Holiday Work Approval", "holiday_work_approval"),
        ("Holiday Title", "holiday_title"),
        ("Verified By", "verified_by"),
        ("Status", "status"),
    ]

    return add_detail_sheet(
        wb,
        "Field Attendance Details",
        columns,
        field_attendance_detail_rows(attendance_logs),
    )


def create_holiday_work_detail_sheet(wb, holiday_work_requests):
    columns = [
        ("Sl No", "sl_no"),
        ("Employee Name", "employee_name"),
        ("Emp Code", "employee_code"),
        ("Department", "department"),
        ("Designation", "designation"),
        ("Team Leader", "team_leader"),
        ("Reporting Officer", "reporting_officer"),
        ("Holiday Date", "date"),
        ("Holiday Title", "holiday_title"),
        ("Holiday Type", "holiday_type"),
        ("Reason", "reason"),
        ("Work Location", "work_location"),
        ("Proof Photo URL", "proof_photo_url"),
        ("Location", "location"),
        ("Map Link", "map_url"),
        ("Approval Stage", "approval_stage"),
        ("Status", "status"),
        ("Decided By", "decided_by"),
        ("Decided At", "decided_at"),
        ("Created At", "created_at"),
    ]

    return add_detail_sheet(
        wb,
        "Holiday Work Requests",
        columns,
        holiday_work_detail_rows(holiday_work_requests),
    )


def create_compoff_detail_sheet(wb, compoff_credits):
    columns = [
        ("Sl No", "sl_no"),
        ("Employee Name", "employee_name"),
        ("Emp Code", "employee_code"),
        ("Department", "department"),
        ("Designation", "designation"),
        ("Team Leader", "team_leader"),
        ("Reporting Officer", "reporting_officer"),
        ("Earned Date", "earned_date"),
        ("Claim From Date", "claim_from_date"),
        ("Expiry Date", "expiry_date"),
        ("Claim Date", "claim_date"),
        ("Holiday Title", "holiday_title"),
        ("Holiday Work Request ID", "holiday_work_request_id"),
        ("Attendance Log ID", "attendance_log_id"),
        ("Leave Request ID", "leave_request_id"),
        ("Status", "status"),
    ]

    return add_detail_sheet(
        wb,
        "Comp-Off Details",
        columns,
        compoff_detail_rows(compoff_credits),
    )


def create_guidelines_sheet(wb):
    ws = wb.active
    ws.title = "Guidelines"

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 12

    ws["B3"] = "Attendance Code Guidelines"
    ws["B3"].font = Font(name="Calibri", size=14, bold=True)
    ws["B3"].fill = TITLE_FILL
    ws["B3"].alignment = Alignment(horizontal="center")
    ws["B3"].border = THIN_BORDER
    ws.merge_cells("B3:C3")

    start_row = 4

    for index, (label, code) in enumerate(GUIDELINE_ROWS, start=start_row):
        ws.cell(index, 2).value = label
        ws.cell(index, 3).value = code
        apply_cell_style(ws.cell(index, 2), align="left")
        apply_cell_style(ws.cell(index, 3), bold=True)

    return ws


def create_attendance_sheet(
    wb,
    employees,
    attendance_logs,
    leave_requests,
    holidays,
    dates,
    organisation_name="",
    organisation_code="",
    state_name="",
    period_label="",
):
    sheet_title = safe_sheet_title(organisation_code or organisation_name or "Attendance")
    ws = wb.create_sheet(sheet_title)

    date_count = len(dates)
    first_day_col = 7
    last_day_col = first_day_col + date_count - 1
    summary_start_col = last_day_col + 1
    final_col = summary_start_col + len(SUMMARY_COLUMNS) - 1

    final_col_letter = get_column_letter(final_col)

    display_org_name = (
        ORG_FULL_NAMES.get(normalize_text(organisation_code).upper())
        or normalize_text(organisation_name)
        or normalize_text(organisation_code)
        or "Organisation"
    )

    period_title = period_label or f"Attendance Record for the Month of {month_name(dates)}"
    if organisation_code:
        period_title = f"{period_title} ({organisation_code})"

    location_title = state_name or "All Locations"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=final_col)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=final_col)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=final_col)

    ws["A1"] = display_org_name
    ws["A2"] = period_title
    ws["A3"] = location_title

    apply_cell_style(ws["A1"], fill=TITLE_FILL, bold=True, size=14)
    apply_cell_style(ws["A2"], fill=SUBTITLE_FILL, bold=True, size=12)
    apply_cell_style(ws["A3"], fill=SUBTITLE_FILL, bold=True, size=12)

    for row in [1, 2, 3]:
        for col in range(1, final_col + 1):
            cell = ws.cell(row, col)
            cell.border = THIN_BORDER
            if row == 1:
                cell.fill = TITLE_FILL
            else:
                cell.fill = SUBTITLE_FILL

    for index, header in enumerate(BASE_COLUMNS, start=1):
        ws.cell(4, index).value = header
        ws.merge_cells(start_row=4, start_column=index, end_row=5, end_column=index)
        apply_cell_style(ws.cell(4, index), fill=HEADER_FILL, bold=True)

    for offset, current_date in enumerate(dates):
        col = first_day_col + offset
        ws.cell(4, col).value = current_date.day
        ws.cell(5, col).value = current_date.strftime("%a").upper()

        header_fill = WEEKEND_FILL if current_date.weekday() == 6 else HEADER_FILL
        apply_cell_style(ws.cell(4, col), fill=header_fill, bold=True)
        apply_cell_style(ws.cell(5, col), fill=header_fill, bold=True)

    for offset, (label, _) in enumerate(SUMMARY_COLUMNS):
        col = summary_start_col + offset
        ws.cell(4, col).value = label
        ws.merge_cells(start_row=4, start_column=col, end_row=5, end_column=col)
        apply_cell_style(ws.cell(4, col), fill=HEADER_FILL, bold=True)

    attendance_lookup = build_attendance_lookup(
        attendance_logs,
        employees=employees,
    )
    leave_lookup = build_leave_lookup(
        leave_requests,
        employees=employees,
    )
    holiday_lookup = build_holiday_lookup(holidays)

    sorted_employees = sorted(
        employees or [],
        key=lambda item: (
            normalize_text(item.get("state") or item.get("branch") or item.get("location")).lower(),
            employee_name(item).lower(),
        ),
    )

    start_row = 6

    for row_offset, employee in enumerate(sorted_employees):
        row = start_row + row_offset

        ws.cell(row, 1).value = row_offset + 1
        ws.cell(row, 2).value = employee_name(employee)
        ws.cell(row, 3).value = employee_designation(employee)
        ws.cell(row, 4).value = employee_project(employee)
        ws.cell(row, 5).value = employee_location(employee)
        ws.cell(row, 6).value = employee_code(employee)

        for col in range(1, 7):
            apply_cell_style(ws.cell(row, col), align="left" if col in {2, 3, 4, 5} else "center")

        row_codes = []

        for offset, current_date in enumerate(dates):
            col = first_day_col + offset
            code = code_for_employee_date(
                employee,
                current_date,
                attendance_lookup,
                leave_lookup,
                holiday_lookup,
            )
            row_codes.append(code)
            ws.cell(row, col).value = code
            style_status_cell(ws.cell(row, col), code)

        cl_total = count_code(row_codes, "CL") + (count_code(row_codes, "CLH") * 0.5)
        el_total = count_code(row_codes, "EL") + (count_code(row_codes, "ELH") * 0.5)
        lwp_total = count_code(row_codes, "LWP") + (count_code(row_codes, "LWPH") * 0.5)
        half_day_total = count_half_day_codes(row_codes)

        ws.cell(row, summary_start_col).value = cl_total
        ws.cell(row, summary_start_col + 1).value = el_total
        ws.cell(row, summary_start_col + 2).value = lwp_total
        ws.cell(row, summary_start_col + 3).value = half_day_total
        ws.cell(row, summary_start_col + 4).value = normalize_text(employee.get("remarks") or "")

        for col in range(summary_start_col, final_col + 1):
            apply_cell_style(ws.cell(row, col))

    if not sorted_employees:
        ws.cell(start_row, 1).value = "No employees found for selected filters."
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=final_col)
        apply_cell_style(ws.cell(start_row, 1), align="center", bold=True)

    total_row = start_row + max(len(sorted_employees), 1) + 1
    ws.cell(total_row, 1).value = "Prepared By"
    ws.cell(total_row, 4).value = "Checked By"
    ws.cell(total_row, 7).value = "Approved By"

    for col in range(1, final_col + 1):
        apply_cell_style(ws.cell(total_row, col), fill=SECTION_FILL, bold=col in {1, 4, 7})

    column_widths = {
        1: 8,
        2: 28,
        3: 24,
        4: 18,
        5: 20,
        6: 14,
    }

    for col, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    for col in range(first_day_col, summary_start_col):
        ws.column_dimensions[get_column_letter(col)].width = 5

    ws.column_dimensions[get_column_letter(summary_start_col)].width = 12
    ws.column_dimensions[get_column_letter(summary_start_col + 1)].width = 12
    ws.column_dimensions[get_column_letter(summary_start_col + 2)].width = 10
    ws.column_dimensions[get_column_letter(summary_start_col + 3)].width = 10
    ws.column_dimensions[get_column_letter(summary_start_col + 4)].width = 24

    for row in range(1, total_row + 1):
        ws.row_dimensions[row].height = 22

    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 22

    ws.freeze_panes = "G6"
    ws.auto_filter.ref = f"A5:{final_col_letter}{total_row - 2}"

    page_setup = ws.page_setup
    page_setup.orientation = "landscape"
    page_setup.fitToWidth = 1
    page_setup.fitToHeight = 0

    ws.sheet_properties.pageSetUpPr.fitToPage = True

    return ws


def build_attendance_workbook(
    employees=None,
    attendance_logs=None,
    leave_requests=None,
    holidays=None,
    holiday_work_requests=None,
    compoff_credits=None,
    period="month",
    year=None,
    month=None,
    date_value=None,
    week_start=None,
    week_end=None,
    organisation_name="",
    organisation_code="",
    state_name="",
):
    dates = build_period_dates(
        period=period,
        year=year,
        month=month,
        date_value=date_value,
        week_start=week_start,
        week_end=week_end,
    )

    wb = Workbook()
    create_guidelines_sheet(wb)

    period_label = ""

    period_key = normalize_text(period).lower()

    if period_key == "day":
        period_label = f"Attendance Record for {dates[0].strftime('%d %B %Y')}"
    elif period_key == "week":
        period_label = f"Attendance Record from {dates[0].strftime('%d %B %Y')} to {dates[-1].strftime('%d %B %Y')}"
    elif period_key == "year":
        period_label = f"Attendance Record for the Year {dates[0].year}"
    else:
        period_label = f"Attendance Record for the Month of {month_name(dates)}"

    create_attendance_sheet(
        wb=wb,
        employees=employees or [],
        attendance_logs=attendance_logs or [],
        leave_requests=leave_requests or [],
        holidays=holidays or [],
        dates=dates,
        organisation_name=organisation_name,
        organisation_code=organisation_code,
        state_name=state_name,
        period_label=period_label,
    )

    exception_rows = build_attendance_exception_rows(
        employees=employees or [],
        attendance_logs=attendance_logs or [],
        leave_requests=leave_requests or [],
        holidays=holidays or [],
        dates=dates,
    )

    report_metadata = [
        f"Period: {report_period_label(dates, period_key)}",
        (
            "Organisation / Entity: "
            f"{normalize_text(organisation_name) or 'All Organisations'}"
            f"{f' ({normalize_text(organisation_code)})' if normalize_text(organisation_code) else ''}"
        ),
        f"State: {normalize_text(state_name) or 'All States'}",
    ]

    add_professional_table_sheet(
        wb,
        "Absent Employees",
        ABSENT_REPORT_COLUMNS,
        exception_rows["absent"],
        metadata=[
            *report_metadata,
            f"Total absent records: {exception_rows['summary']['absent_records']}",
        ],
        sheet_title="Absent Employees",
    )

    add_professional_table_sheet(
        wb,
        "Employees Without Check-Out",
        MISSING_CHECKOUT_REPORT_COLUMNS,
        exception_rows["missing_checkout"],
        metadata=[
            *report_metadata,
            (
                "Total missing-checkout records: "
                f"{exception_rows['summary']['missing_checkout_records']}"
            ),
        ],
        sheet_title="Missing Check-Out",
    )

    create_field_attendance_detail_sheet(wb, attendance_logs or [])
    create_holiday_work_detail_sheet(wb, holiday_work_requests or [])
    create_compoff_detail_sheet(wb, compoff_credits or [])

    return wb


def workbook_to_bytes(workbook):
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def build_attendance_excel_file(
    employees=None,
    attendance_logs=None,
    leave_requests=None,
    holidays=None,
    holiday_work_requests=None,
    compoff_credits=None,
    period="month",
    year=None,
    month=None,
    date_value=None,
    week_start=None,
    week_end=None,
    organisation_name="",
    organisation_code="",
    state_name="",
):
    workbook = build_attendance_workbook(
        employees=employees,
        attendance_logs=attendance_logs,
        leave_requests=leave_requests,
        holidays=holidays,
        holiday_work_requests=holiday_work_requests,
        compoff_credits=compoff_credits,
        period=period,
        year=year,
        month=month,
        date_value=date_value,
        week_start=week_start,
        week_end=week_end,
        organisation_name=organisation_name,
        organisation_code=organisation_code,
        state_name=state_name,
    )

    return workbook_to_bytes(workbook)


def build_attendance_exception_workbook(
    employees=None,
    attendance_logs=None,
    leave_requests=None,
    holidays=None,
    period="month",
    year=None,
    month=None,
    date_value=None,
    week_start=None,
    week_end=None,
    organisation_name="",
    organisation_code="",
    state_name="",
    report_type="absent",
    today_reference=None,
):
    dates = build_period_dates(
        period=period,
        year=year,
        month=month,
        date_value=date_value,
        week_start=week_start,
        week_end=week_end,
    )
    exception_rows = build_attendance_exception_rows(
        employees=employees or [],
        attendance_logs=attendance_logs or [],
        leave_requests=leave_requests or [],
        holidays=holidays or [],
        dates=dates,
        today_reference=today_reference,
    )

    report_type = normalize_text(report_type).lower().replace("-", "_")

    if report_type not in {"absent", "missing_checkout", "exceptions"}:
        report_type = "absent"

    wb = Workbook()
    metadata = [
        f"Period: {report_period_label(dates, period)}",
        (
            "Organisation / Entity: "
            f"{normalize_text(organisation_name) or 'All Organisations'}"
            f"{f' ({normalize_text(organisation_code)})' if normalize_text(organisation_code) else ''}"
        ),
        f"State: {normalize_text(state_name) or 'All States'}",
        f"Generated: {datetime.now().strftime('%d %B %Y, %I:%M %p')}",
    ]

    if report_type in {"absent", "exceptions"}:
        add_professional_table_sheet(
            wb,
            "Absent Employees",
            ABSENT_REPORT_COLUMNS,
            exception_rows["absent"],
            metadata=[
                *metadata,
                f"Total absent records: {exception_rows['summary']['absent_records']}",
            ],
            sheet_title="Absent Employees",
        )

    if report_type in {"missing_checkout", "exceptions"}:
        add_professional_table_sheet(
            wb,
            "Employees Without Check-Out",
            MISSING_CHECKOUT_REPORT_COLUMNS,
            exception_rows["missing_checkout"],
            metadata=[
                *metadata,
                (
                    "Total missing-checkout records: "
                    f"{exception_rows['summary']['missing_checkout_records']}"
                ),
            ],
            sheet_title="Missing Check-Out",
        )

    return wb


def build_attendance_exception_excel_file(**kwargs):
    return workbook_to_bytes(build_attendance_exception_workbook(**kwargs))


EMPLOYEE_MASTER_COLUMNS = [
    ("Sl No", "sl_no"),
    ("Employee Code", "employee_code"),
    ("Employee Name", "employee_name"),
    ("Official Email", "official_email"),
    ("Personal Email", "personal_email"),
    ("Phone", "phone"),
    ("Organisation Code", "organisation_code"),
    ("Organisation / Entity", "organisation"),
    ("State", "state"),
    ("Department", "department"),
    ("Designation", "designation"),
    ("Role", "role"),
    ("Employee Type", "employee_type"),
    ("Job Type", "job_type"),
    ("Shift", "shift"),
    ("Joining Date", "joining_date"),
    ("Employment Status", "status"),
    ("Team Leader", "team_leader"),
    ("Reporting Officer", "reporting_officer"),
    ("Gender", "gender"),
    ("Date of Birth", "date_of_birth"),
    ("Blood Group", "blood_group"),
    ("PAN No", "pan_no"),
    ("Aadhaar No", "aadhar_no"),
    ("UAN No", "uan_no"),
    ("ESIC IP", "esic_ip"),
    ("Gross Salary", "gross_salary"),
    ("Payment Mode", "payment_mode"),
    ("Address", "address"),
    ("Last Working Date", "last_working_date"),
    ("Exit Type", "exit_type"),
    ("Exit Reason", "exit_reason"),
]


def employee_master_rows(employees):
    rows = []
    seen_employee_ids = set()

    for employee in employees or []:
        stable_id = normalize_text(employee.get("_id") or employee.get("id"))

        if stable_id and stable_id in seen_employee_ids:
            continue

        if stable_id:
            seen_employee_ids.add(stable_id)

        official_email = normalize_text(
            employee.get("official_email") or employee.get("email")
        )
        personal_email = normalize_text(
            employee.get("personal_email") or employee.get("alternate_email")
        )
        status = normalize_text(
            employee.get("employment_status") or employee.get("status") or "active"
        )

        rows.append({
            "employee_code": employee_code(employee),
            "employee_name": employee_name(employee),
            "official_email": official_email,
            "personal_email": personal_email,
            "phone": normalize_text(employee.get("phone") or employee.get("mobile")),
            "organisation_code": employee_organisation_code(employee),
            "organisation": employee_organisation_name(employee),
            "state": employee_state(employee),
            "department": employee_department(employee),
            "designation": employee_designation(employee),
            "role": normalize_text(employee.get("role") or "employee"),
            "employee_type": normalize_text(employee.get("employee_type")),
            "job_type": normalize_text(employee.get("job_type")),
            "shift": normalize_text(employee.get("shift")),
            "joining_date": normalize_text(
                employee.get("joining_date") or employee.get("date_of_joining")
            ),
            "status": status,
            "team_leader": normalize_text(employee.get("team_leader_name")),
            "reporting_officer": normalize_text(employee.get("reporting_officer_name")),
            "gender": normalize_text(employee.get("gender")),
            "date_of_birth": normalize_text(
                employee.get("date_of_birth") or employee.get("dob")
            ),
            "blood_group": normalize_text(employee.get("blood_group")),
            "pan_no": normalize_text(employee.get("pan_no") or employee.get("pan")),
            "aadhar_no": normalize_text(
                employee.get("aadhar_no") or employee.get("aadhaar_no")
            ),
            "uan_no": normalize_text(
                employee.get("employee_uan_no") or employee.get("uan_no")
            ),
            "esic_ip": normalize_text(
                employee.get("employee_esic_ip") or employee.get("esic_ip")
            ),
            "gross_salary": employee.get("gross_salary", employee.get("salary", "")),
            "payment_mode": normalize_text(employee.get("payment_mode")),
            "address": normalize_text(
                employee.get("address") or employee.get("current_address")
            ),
            "last_working_date": normalize_text(
                employee.get("last_working_date") or employee.get("resignation_date")
            ),
            "exit_type": normalize_text(employee.get("exit_type")),
            "exit_reason": normalize_text(
                employee.get("resignation_reason") or employee.get("exit_reason")
            ),
        })

    rows.sort(
        key=lambda row: (
            row["organisation_code"].lower(),
            row["state"].lower(),
            row["employee_name"].lower(),
        )
    )

    for index, row in enumerate(rows, start=1):
        row["sl_no"] = index

    return rows


def build_employee_master_workbook(
    employees=None,
    tenant_name="",
    organisation_name="",
    organisation_code="",
    state_name="",
    employee_scope="active",
    generated_by="",
    applied_filters=None,
):
    rows = employee_master_rows(employees or [])
    wb = Workbook()
    organisation_label = normalize_text(organisation_name) or "All Organisations"
    organisation_code = normalize_text(organisation_code)

    if organisation_code:
        organisation_label = f"{organisation_label} ({organisation_code})"

    metadata = [
        f"Tenant / Company: {normalize_text(tenant_name) or 'Current Tenant'}",
        f"Organisation / Entity: {organisation_label}",
        f"State: {normalize_text(state_name) or 'All States'}",
        f"Employee Scope: {normalize_text(employee_scope).title() or 'Active'}",
        f"Total Employees: {len(rows)}",
        f"Generated: {datetime.now().strftime('%d %B %Y, %I:%M %p')}",
    ]

    filter_text = ", ".join(
        f"{normalize_text(key).replace('_', ' ').title()}: {normalize_text(value)}"
        for key, value in (applied_filters or {}).items()
        if normalize_text(value)
    )

    if filter_text:
        metadata.append(f"Applied Filters: {filter_text}")

    if normalize_text(generated_by):
        metadata.append(f"Generated By: {normalize_text(generated_by)}")

    add_professional_table_sheet(
        wb,
        "Employee Master Report",
        EMPLOYEE_MASTER_COLUMNS,
        rows,
        metadata=metadata,
        sheet_title="Employee Master",
    )

    ws = wb["Employee Master"]
    status_column = next(
        index
        for index, (_, key) in enumerate(EMPLOYEE_MASTER_COLUMNS, start=1)
        if key == "status"
    )
    header_row = ws.auto_filter.ref.split(":", 1)[0]
    header_row_number = int(re.sub(r"[^0-9]", "", header_row) or 1)

    for row_number in range(header_row_number + 1, header_row_number + len(rows) + 1):
        status_cell = ws.cell(row_number, status_column)
        status = normalize_text(status_cell.value).lower()

        if status in {"active", "confirmed", "probation"}:
            status_cell.fill = WEEKEND_FILL
        elif status in {"resigned", "left", "terminated", "inactive", "retired"}:
            status_cell.fill = ABSENT_FILL

    return wb


def build_employee_master_excel_file(**kwargs):
    return workbook_to_bytes(build_employee_master_workbook(**kwargs))


def sanitize_excel_filename(filename):
    filename = normalize_text(filename) or "report.xlsx"

    for char in [" ", "/", "\\", ":", "*", "?", '"', "<", ">", "|", "(", ")"]:
        filename = filename.replace(char, "_")

    while "__" in filename:
        filename = filename.replace("__", "_")

    return filename


def build_employee_master_excel_filename(
    organisation_code="",
    organisation_name="",
    state_name="",
    employee_scope="active",
):
    organisation_part = (
        normalize_text(organisation_code)
        or normalize_text(organisation_name)
        or "All_Organisations"
    )
    state_part = normalize_text(state_name) or "All_States"
    scope_part = normalize_text(employee_scope).title() or "Active"
    date_part = datetime.now().strftime("%Y_%m_%d")

    return sanitize_excel_filename(
        f"{organisation_part}_{state_part}_{scope_part}_Employees_{date_part}.xlsx"
    )


def build_attendance_exception_excel_filename(
    report_type="absent",
    organisation_code="",
    organisation_name="",
    state_name="",
    period="month",
    year=None,
    month=None,
    date_value=None,
    week_start=None,
    week_end=None,
):
    dates = build_period_dates(
        period=period,
        year=year,
        month=month,
        date_value=date_value,
        week_start=week_start,
        week_end=week_end,
    )
    report_key = normalize_text(report_type).lower().replace("-", "_")
    report_label = {
        "absent": "Absent_Employees",
        "missing_checkout": "Missing_Checkout",
        "exceptions": "Attendance_Exceptions",
    }.get(report_key, "Absent_Employees")
    organisation_part = (
        normalize_text(organisation_code)
        or normalize_text(organisation_name)
        or "All_Organisations"
    )
    state_part = normalize_text(state_name) or "All_States"
    period_key = normalize_text(period).lower()

    if period_key == "day":
        period_part = dates[0].strftime("%Y_%m_%d")
    elif period_key == "week":
        period_part = (
            f"{dates[0].strftime('%Y_%m_%d')}_to_"
            f"{dates[-1].strftime('%Y_%m_%d')}"
        )
    elif period_key == "year":
        period_part = str(dates[0].year)
    else:
        period_part = dates[0].strftime("%Y_%m")

    return sanitize_excel_filename(
        f"{organisation_part}_{state_part}_{report_label}_{period_part}.xlsx"
    )


def build_attendance_excel_filename(
    organisation_code="",
    organisation_name="",
    state_name="",
    period="month",
    year=None,
    month=None,
    date_value=None,
    week_start=None,
    week_end=None,
):
    dates = build_period_dates(
        period=period,
        year=year,
        month=month,
        date_value=date_value,
        week_start=week_start,
        week_end=week_end,
    )

    org_part = (
        normalize_text(organisation_code)
        or normalize_text(organisation_name)
        or "Organisation"
    )

    state_part = normalize_text(state_name) or "AllStates"

    period_key = normalize_text(period).lower()

    if period_key == "day":
        period_part = dates[0].strftime("%d_%b_%Y")
    elif period_key == "week":
        period_part = f"{dates[0].strftime('%d_%b_%Y')}_to_{dates[-1].strftime('%d_%b_%Y')}"
    elif period_key == "year":
        period_part = str(dates[0].year)
    else:
        period_part = dates[0].strftime("%B_%Y")

    return sanitize_excel_filename(
        f"{org_part}_{state_part}_Attendance_{period_part}.xlsx"
    )
